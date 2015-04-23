from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
wb = load_workbook('sample.xlsx')

ws = wb.active

sum = 0

for row in ws.iter_rows('A2:D6'):
	# print row[2].value
	if row[2].value == 'ayam':
		sum += row[3].value
print sum